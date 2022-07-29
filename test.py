from openpyxl import Workbook,  load_workbook
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter
from PIL import Image
wb = load_workbook('list.xlsx')
ws = wb['Sheet1']
image_loader = SheetImageLoader(ws)
print(ws['A1'])

myCell = ws['A5']
print(myCell.row, myCell.column_letter)
