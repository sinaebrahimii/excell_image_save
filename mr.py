import code
from html.entities import codepoint2name
from logging import exception
from openpyxl import Workbook,  load_workbook
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter
from PIL import Image
wb = load_workbook('list.xlsx')
ws = wb['Sheet1']
image_loader = SheetImageLoader(ws)
codeMelliColumn = ''
pictureColumn = ''
codeMelli = ''
img = ''
# find the code and picture column
for cell in ws[1]:
    if cell.value == 'شماره ملی':
        codeMelliColumn = cell.column_letter

for cell in ws[1]:
    if cell.value == 'عکس4*3':
        pictureColumn = cell.column_letter
# -------------------------------------------
print(pictureColumn)
# iterate first and last row
for row in range(1, ws.max_row + 1):
    if ws[f'{codeMelliColumn}{row}'].value == None or ws[f'{codeMelliColumn}{row}'].value == 'شماره ملی':
        pass
    else:
        codeMelli = ws[f'{codeMelliColumn}{row}'].value
    if image_loader.image_in(f"{pictureColumn}{row}"):
        print('yes', f'{pictureColumn}{row}')
        img = image_loader.get(f"{pictureColumn}{row}")
        img.save(f"imgaes/{codeMelli}.jpg")
